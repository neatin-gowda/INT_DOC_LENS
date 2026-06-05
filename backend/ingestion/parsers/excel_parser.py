"""
Excel Spreadsheet (.xlsx, .xlsm, .xlsb, .xls) parser.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from ...models import Block, BlockType
from .utils import (
    _clean,
    _slug,
    _detect_header_band,
    _block,
    _page_for_sequence,
    _row_payload,
    _detect_stable_key,
    _row_text,
)

def _sheet_rows_from_openpyxl(source_path: Path) -> Iterable[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(source_path), data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = []
            for raw_row in sheet.iter_rows(values_only=True):
                row = [_clean(value) for value in raw_row]
                if any(row):
                    rows.append(row)
            yield sheet.title, rows
    finally:
        workbook.close()

def _sheet_rows_from_xls(source_path: Path) -> Iterable[tuple[str, list[list[str]]]]:
    try:
        import xlrd
    except Exception:
        return []

    workbook = xlrd.open_workbook(str(source_path))
    out = []
    for sheet in workbook.sheets():
        rows = []
        for ri in range(sheet.nrows):
            row = [_clean(sheet.cell_value(ri, ci)) for ci in range(sheet.ncols)]
            if any(row):
                rows.append(row)
        out.append((sheet.name, rows))
    return out

def _sheet_rows_from_xlsb(source_path: Path) -> Iterable[tuple[str, list[list[str]]]]:
    try:
        from pyxlsb import open_workbook
    except Exception:
        return []

    out = []
    with open_workbook(str(source_path)) as workbook:
        for sheet_name in workbook.sheets:
            rows = []
            with workbook.get_sheet(sheet_name) as sheet:
                for raw_row in sheet.rows():
                    row = [_clean(cell.v if cell is not None else "") for cell in raw_row]
                    if any(row):
                        rows.append(row)
            out.append((sheet_name, rows))
    return out

def _extract_spreadsheet(source_path: Path) -> list[Block]:
    ext = source_path.suffix.lower()

    try:
        if ext in {".xlsx", ".xlsm"}:
            sheets = list(_sheet_rows_from_openpyxl(source_path))
        elif ext == ".xlsb":
            sheets = list(_sheet_rows_from_xlsb(source_path))
        elif ext == ".xls":
            sheets = list(_sheet_rows_from_xls(source_path))
        else:
            from .csv_parser import _sheet_rows_from_csv
            sheets = list(_sheet_rows_from_csv(source_path))
    except Exception:
        return []

    blocks: list[Block] = []
    seq = 0

    for sheet_name, rows in sheets:
        rows = [row for row in rows if any(_clean(cell) for cell in row)]
        if not rows:
            continue

        n_cols = max(len(row) for row in rows)
        normalized_rows = [row + [""] * (n_cols - len(row)) for row in rows]

        header, body_rows, header_rows, header_index, header_strategy = _detect_header_band(normalized_rows, n_cols)

        sheet_slug = _slug(sheet_name, f"sheet_{seq}")
        section = _block(
            block_type=BlockType.SECTION,
            path=f"/{sheet_slug}",
            text=sheet_name,
            payload={
                "heading": sheet_name,
                "source_format": ext.lstrip("."),
                "sheet_name": sheet_name,
            },
            sequence=seq,
            page_number=_page_for_sequence(seq, 55),
        )
        blocks.append(section)
        seq += 1

        table_payload = {
            "header": header,
            "header_rows": header_rows,
            "header_row_count": len(header_rows),
            "header_index": header_index,
            "header_strategy": header_strategy,
            "rows": body_rows,
            "spans_pages": [_page_for_sequence(seq, 55)],
            "table_title": sheet_name,
            "table_context": f"Sheet: {sheet_name}",
            "source_format": ext.lstrip("."),
            "sheet_name": sheet_name,
        }
        table = _block(
            parent_id=section.id,
            block_type=BlockType.TABLE,
            path=f"/{sheet_slug}/table_{seq}",
            text=sheet_name,
            payload=table_payload,
            sequence=seq,
            page_number=_page_for_sequence(seq, 55),
        )
        blocks.append(table)
        seq += 1

        for ri, row in enumerate(body_rows):
            payload = _row_payload(header, row)
            payload.update(
                {
                    "__row_index__": ri,
                    "__table_title__": sheet_name,
                    "__pages__": [_page_for_sequence(seq, 55)],
                    "source_format": ext.lstrip("."),
                    "sheet_name": sheet_name,
                }
            )
            text = _row_text(payload)
            block = _block(
                parent_id=table.id,
                block_type=BlockType.TABLE_ROW,
                path=f"{table.path}/row_{ri}",
                text=text,
                payload=payload,
                sequence=seq,
                page_number=_page_for_sequence(seq, 55),
                stable_key=_detect_stable_key(row, header),
            )
            blocks.append(block)
            seq += 1

    return blocks
